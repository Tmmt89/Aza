#!/usr/bin/env python3
"""Добавляет в каталог расписаний города Северного Кавказа, по которым
исследование собрало годовые официальные данные.

Источники (см. docs/PLAN-prayer-schedules.md):
  Махачкала  — Муфтият Республики Дагестан, годовой JSON;
  Нальчик    — ДУМ Кабардино-Балкарской Республики, годовой PDF;
  Грозный    — расписание ДУМ ЧР, годовой XLSX (путь через --grozny).

Скрипт ничего не выдумывает: если день не разобрался или времена идут не
по возрастанию, город не добавляется вовсе. Лучше расчёт с честной
подписью, чем таблица с дырами.

Запуск:
  python3 Tools/import-caucasus-schedules.py <папка-исследования> [--write]
"""
import hashlib
import io
import json
import os
import re
import sys

YEAR = 2026
MONTHS = {
    "ЯНВАРЬ": 1, "ФЕВРАЛЬ": 2, "МАРТ": 3, "АПРЕЛЬ": 4, "МАЙ": 5, "ИЮНЬ": 6,
    "ИЮЛЬ": 7, "АВГУСТ": 8, "СЕНТЯБРЬ": 9, "ОКТЯБРЬ": 10, "НОЯБРЬ": 11,
    "ДЕКАБРЬ": 12,
}
DAYS_IN_MONTH = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]


def sha256(path):
    with open(path, "rb") as f:
        return hashlib.sha256(f.read()).hexdigest()


def valid_day(times):
    """Шесть значений ЧЧ:ММ строго по возрастанию."""
    if len(times) != 6:
        return False
    minutes = []
    for value in times:
        match = re.fullmatch(r"([0-2]\d):([0-5]\d)", value)
        if not match:
            return False
        hour, minute = int(match.group(1)), int(match.group(2))
        if hour > 23:
            return False
        minutes.append(hour * 60 + minute)
    return minutes == sorted(minutes) and len(set(minutes)) == 6


# Часовые пояса берутся отсюда, а не подставляются московским: для
# Екатеринбурга, Самары и Волгограда это сдвиг на час-два, то есть
# заведомо неверные времена намаза.
TIME_ZONES = {
    "Грозный": "Europe/Moscow",
    "Махачкала": "Europe/Moscow",
    "Нальчик": "Europe/Moscow",
    "Казань": "Europe/Moscow",
    "Москва": "Europe/Moscow",
    "Ростов-на-Дону": "Europe/Moscow",
    "Самара": "Europe/Samara",
    "Волгоград": "Europe/Volgograd",
    "Екатеринбург": "Asia/Yekaterinburg",
}


def city_entry(city_id, name, source, days):
    zone = TIME_ZONES.get(name)
    if zone is None:
        raise ValueError(f"{name}: часовой пояс неизвестен — добавьте его в TIME_ZONES")
    return {
        "id": city_id,
        "name": name,
        "timeZone": zone,
        "coverageStatus": "complete",
        "coverageStart": f"{YEAR}-01-01",
        "coverageEnd": f"{YEAR}-12-31",
        "releaseStatus": "ownerApproved",
        "source": source,
        "days": days,
    }


def read_makhachkala(root):
    path = os.path.join(root, "work/prayer-research/official/dum-dagestan-makhachkala.json")
    raw = json.load(io.open(path, encoding="utf-8"))
    if len(raw) != 365:
        raise ValueError(f"Махачкала: ожидалось 365 дней, получено {len(raw)}")
    days, day_of_year = [], 0
    for month_index, count in enumerate(DAYS_IN_MONTH, start=1):
        for day in range(1, count + 1):
            row = raw[day_of_year]
            day_of_year += 1
            times = [row["namaz_1"], row["voshod"], row["namaz_2"],
                     row["namaz_3"], row["namaz_4"], row["namaz_5"]]
            if not valid_day(times):
                raise ValueError(f"Махачкала: день {month_index:02d}-{day:02d} не прошёл проверку: {times}")
            days.append({"date": f"{YEAR}-{month_index:02d}-{day:02d}", "times": times})
    return city_entry("махачкала", "Махачкала", {
        "name": "Муфтият Республики Дагестан",
        "url": "https://muftiyatrd.ru/json/namaz/mahashkala.json",
        "sha256": sha256(path),
    }, days)


def read_govzalla_xlsx(path, source_name):
    """Годовой XLSX формата «Ламазан хенаш»: лист «Весь год», колонки
    Дата, День, Месяц и шесть времён. Название города берётся из
    заголовка, чтобы оно не разъезжалось с содержимым файла.

    Внимание: такие файлы бывают и настоящей таблицей муфтията, и
    расчётом. Отличать — по фиксированному зухру (см. документацию);
    скрипт этого НЕ решает, решение принимает человек.
    """
    import openpyxl
    book = openpyxl.load_workbook(path, data_only=True)
    sheet = book["Весь год"]
    title = str(sheet.cell(1, 1).value or "")
    parts = [p.strip() for p in title.split("•")]
    if len(parts) < 2 or not parts[1]:
        raise ValueError(f"{os.path.basename(path)}: не разобрал город из заголовка «{title}»")
    name = parts[1]
    found = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not isinstance(row[0], str):
            continue
        parts = row[0].split(".")
        if len(parts) != 3 or not all(p.isdigit() for p in parts):
            continue
        day, month, year = (int(p) for p in parts)
        if year != YEAR:
            continue
        times = [str(value)[:5] for value in row[3:9] if value is not None]
        if not valid_day(times):
            raise ValueError(f"Грозный: день {row[0]} не прошёл проверку: {times}")
        # Повтор даты — испорченный файл, а не «возьмём последнюю строку»:
        # молчаливая перезапись выбирала времена порядком строк в XLSX.
        if (month, day) in found:
            raise ValueError(f"{os.path.basename(path)}: дата {row[0]} встречается дважды")
        found[(month, day)] = times

    days = []
    for month_index, count in enumerate(DAYS_IN_MONTH, start=1):
        for day in range(1, count + 1):
            times = found.get((month_index, day))
            if times is None:
                raise ValueError(f"{name}: нет дня {month_index:02d}-{day:02d}, разобрано {len(found)}")
            days.append({"date": f"{YEAR}-{month_index:02d}-{day:02d}", "times": times})
    return city_entry(name.lower(), name, {
        "name": source_name,
        "url": "https://govzalla.com/ламазан-хенаш-время-молитв",
        "sha256": sha256(path),
    }, days)


def read_nalchik(root):
    path = os.path.join(root, "work/prayer-research/official/kbr-2026.txt")
    pdf = os.path.join(root, "work/prayer-research/official/kbr-2026.pdf")
    text = io.open(path, encoding="utf-8").read()
    month = None
    found = {}
    for line in text.splitlines():
        header = re.search(r"ГРАФИК\s+НАМАЗОВ\s+НА\s+([А-ЯЁ]+)\s+2026", line.upper())
        if header:
            month = MONTHS.get(header.group(1))
            continue
        if month is None:
            continue
        row = re.match(r"\s*(\d{1,2})\s+\S+\s+((?:[0-2]?\d:[0-5]\d\s+){5}[0-2]?\d:[0-5]\d)", line)
        if not row:
            continue
        day = int(row.group(1))
        if not 1 <= day <= DAYS_IN_MONTH[month - 1]:
            continue
        times = [t if len(t) == 5 else "0" + t for t in row.group(2).split()]
        if not valid_day(times):
            continue
        found.setdefault((month, day), times)

    days = []
    for month_index, count in enumerate(DAYS_IN_MONTH, start=1):
        for day in range(1, count + 1):
            times = found.get((month_index, day))
            if times is None:
                raise ValueError(f"Нальчик: нет дня {month_index:02d}-{day:02d}, разобрано {len(found)} из 365")
            days.append({"date": f"{YEAR}-{month_index:02d}-{day:02d}", "times": times})
    return city_entry("нальчик", "Нальчик", {
        "name": "ДУМ Кабардино-Балкарской Республики",
        "url": "https://www.kbrdum.ru/8-grafik-namazov",
        "sha256": sha256(pdf),
    }, days)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 2
    root = sys.argv[1]
    write = "--write" in sys.argv
    catalog_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                "Aza/Resources/prayer-schedules-2026.json")
    catalog = json.load(io.open(catalog_path, encoding="utf-8"))
    known = {c["name"] for c in catalog["cities"]}

    # --xlsx ПУТЬ ПОДПИСЬ — годовой файл формата «Ламазан хенаш».
    # --replace разрешает заменить город, уже лежащий в каталоге: без
    # него существующая запись никогда не перезаписывается молча.
    extra = []
    for i, arg in enumerate(sys.argv):
        if arg == "--xlsx":
            if i + 2 >= len(sys.argv):
                sys.exit("--xlsx требует два аргумента: ПУТЬ ПОДПИСЬ")
            extra.append((sys.argv[i + 1], sys.argv[i + 2]))
    replace = "--replace" in sys.argv

    readers = [read_makhachkala, read_nalchik]
    for path, label in extra:
        readers.append(lambda _root, p=path, l=label: read_govzalla_xlsx(p, l))

    added, replaced = [], []
    processed = set()
    for reader in readers:
        try:
            entry = reader(root)
        except (OSError, ValueError, KeyError) as error:
            print("ПРОПУЩЕН:", error)
            continue
        if entry["name"] in processed:
            print("ПРОПУЩЕН повторный источник города:", entry["name"])
            continue
        processed.add(entry["name"])
        if entry["name"] in known:
            if not replace:
                print("уже в каталоге:", entry["name"])
                continue
            replaced.append(entry)
            print(f"ЗАМЕНА: {entry['name']} — {len(entry['days'])} дней, {entry['source']['name']}")
            continue
        added.append(entry)
        print(f"готов: {entry['name']} — {len(entry['days'])} дней, {entry['source']['name']}")

    if not added and not replaced:
        print("нечего добавлять")
        return 1
    if not write:
        print("\nпробный прогон; применить: --write")
        return 0

    if replaced:
        names = {e["name"] for e in replaced}
        catalog["cities"] = [c for c in catalog["cities"] if c["name"] not in names]
        catalog["cities"].extend(replaced)
    catalog["cities"].extend(added)
    catalog["cities"].sort(key=lambda c: c["name"])
    catalog["cityCount"] = len(catalog["cities"])
    catalog["completeCityCount"] = sum(1 for c in catalog["cities"]
                                       if c["coverageStatus"] == "complete")
    catalog["partialCityCount"] = catalog["cityCount"] - catalog["completeCityCount"]
    # Атомарная запись: обрыв на прямом write оставлял усечённый каталог,
    # и приложение теряло ВСЕ расписания разом.
    tmp_path = catalog_path + ".tmp"
    io.open(tmp_path, "w", encoding="utf-8").write(
        json.dumps(catalog, ensure_ascii=False, separators=(",", ":")))
    os.replace(tmp_path, catalog_path)
    print(f"\nзаписано: {catalog['cityCount']} городов")
    return 0


if __name__ == "__main__":
    sys.exit(main())
